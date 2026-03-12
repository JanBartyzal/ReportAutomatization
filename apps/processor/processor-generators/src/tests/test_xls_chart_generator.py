"""Tests for Excel chart_generator module."""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook, load_workbook

from src.generators.xls.service.chart_generator import ChartGenerator
from src.generators.xls.service.excel_renderer import ChartInput, ChartSeriesInput


def _make_workbook_with_chart(chart_data: ChartInput) -> bytes:
    """Create a workbook, add a chart, and return XLSX bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Dashboard"

    generator = ChartGenerator()
    generator.add_chart(ws, chart_data, anchor_row=3, anchor_col=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class TestBarChartCreation:
    """Test bar chart creation."""

    def test_bar_chart(self):
        chart_data = ChartInput(
            chart_type="BAR",
            title="Revenue by Quarter",
            labels=["Q1", "Q2", "Q3", "Q4"],
            series=[
                ChartSeriesInput(name="Revenue", values=[100.0, 150.0, 120.0, 180.0]),
            ],
        )
        xlsx_bytes = _make_workbook_with_chart(chart_data)
        assert len(xlsx_bytes) > 0

        # Verify it's a valid XLSX with a chart
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert len(ws._charts) == 1

    def test_multiple_series_bar(self):
        chart_data = ChartInput(
            chart_type="BAR",
            title="OPEX vs CAPEX",
            labels=["2023", "2024", "2025"],
            series=[
                ChartSeriesInput(name="OPEX", values=[500.0, 480.0, 460.0]),
                ChartSeriesInput(name="CAPEX", values=[200.0, 250.0, 220.0]),
            ],
        )
        xlsx_bytes = _make_workbook_with_chart(chart_data)

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert len(ws._charts) == 1


class TestLineChartCreation:
    """Test line chart creation."""

    def test_line_chart(self):
        chart_data = ChartInput(
            chart_type="LINE",
            title="Monthly Sales Trend",
            labels=["Jan", "Feb", "Mar"],
            series=[
                ChartSeriesInput(name="Sales", values=[10.0, 20.0, 15.0]),
                ChartSeriesInput(name="Costs", values=[8.0, 12.0, 11.0]),
            ],
        )
        xlsx_bytes = _make_workbook_with_chart(chart_data)

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert len(ws._charts) == 1


class TestPieChartCreation:
    """Test pie chart creation."""

    def test_pie_chart(self):
        chart_data = ChartInput(
            chart_type="PIE",
            title="Budget Distribution",
            labels=["IT", "HR", "Marketing"],
            series=[
                ChartSeriesInput(name="Budget", values=[45.0, 25.0, 30.0]),
            ],
        )
        xlsx_bytes = _make_workbook_with_chart(chart_data)

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert len(ws._charts) == 1

    def test_pie_chart_no_series(self):
        chart_data = ChartInput(
            chart_type="PIE",
            title="Empty Pie",
            labels=["A", "B"],
            series=[],
        )
        xlsx_bytes = _make_workbook_with_chart(chart_data)
        assert len(xlsx_bytes) > 0

    def test_unsupported_chart_type_no_error(self):
        """Unsupported chart types log a warning but don't raise."""
        wb = Workbook()
        ws = wb.active

        chart_data = ChartInput(
            chart_type="RADAR",
            labels=["A"],
            series=[ChartSeriesInput(name="S", values=[1.0])],
        )
        generator = ChartGenerator()
        # Should not raise, just log a warning
        generator.add_chart(ws, chart_data, anchor_row=1, anchor_col=1)

        # No chart should have been added
        assert len(ws._charts) == 0
