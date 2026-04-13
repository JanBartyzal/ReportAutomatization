"""Unit tests for SheetUpdater - partial Excel sheet update engine."""

import io
import pytest
from openpyxl import Workbook, load_workbook

from src.generators.xls.service.sheet_updater import SheetUpdater, MAX_INPUT_SIZE_BYTES


@pytest.fixture
def updater():
    return SheetUpdater()


@pytest.fixture
def multi_sheet_workbook():
    """Create a test workbook with 3 sheets: Data, Summary, Charts."""
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"
    ws_data["A1"] = "Original"
    ws_data["A2"] = "Row1"
    ws_data["B1"] = "Column"
    ws_data["B2"] = "Value1"

    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Total"
    ws_summary["B1"] = 42

    ws_charts = wb.create_sheet("Charts")
    ws_charts["A1"] = "ChartPlaceholder"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


class TestSheetUpdater:

    def test_create_new_workbook_when_empty_binary(self, updater):
        result = updater.update_sheet(
            excel_binary=None,
            sheet_name="NewSheet",
            headers=["Col1", "Col2"],
            data_rows=[["a", "b"], ["c", "d"]],
        )

        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        assert "NewSheet" in wb.sheetnames
        assert result.rows_written == 2
        assert result.sheet_name == "NewSheet"
        ws = wb["NewSheet"]
        assert ws.cell(1, 1).value == "Col1"
        assert ws.cell(1, 2).value == "Col2"
        assert ws.cell(2, 1).value == "a"
        assert ws.cell(3, 2).value == "d"

    def test_update_existing_sheet_preserves_other_sheets(self, updater, multi_sheet_workbook):
        result = updater.update_sheet(
            excel_binary=multi_sheet_workbook,
            sheet_name="Data",
            headers=["New1", "New2", "New3"],
            data_rows=[["x", "y", "z"]],
        )

        wb = load_workbook(io.BytesIO(result.xlsx_bytes))

        # Target sheet updated
        ws_data = wb["Data"]
        assert ws_data.cell(1, 1).value == "New1"
        assert ws_data.cell(2, 1).value == "x"

        # Other sheets preserved
        assert "Summary" in wb.sheetnames
        assert "Charts" in wb.sheetnames
        ws_summary = wb["Summary"]
        assert ws_summary["A1"].value == "Total"
        assert ws_summary["B1"].value == 42
        ws_charts = wb["Charts"]
        assert ws_charts["A1"].value == "ChartPlaceholder"

    def test_create_new_sheet_in_existing_workbook(self, updater, multi_sheet_workbook):
        result = updater.update_sheet(
            excel_binary=multi_sheet_workbook,
            sheet_name="Brand New",
            headers=["H1"],
            data_rows=[["v1"], ["v2"]],
        )

        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        assert "Brand New" in wb.sheetnames
        assert "Data" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        assert "Charts" in wb.sheetnames
        assert result.rows_written == 2

    def test_data_types(self, updater):
        result = updater.update_sheet(
            excel_binary=None,
            sheet_name="Types",
            headers=["String", "Number", "Bool"],
            data_rows=[
                ["hello", 3.14, True],
                ["world", 42, False],
            ],
        )

        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        ws = wb["Types"]
        assert ws.cell(2, 1).value == "hello"
        assert ws.cell(2, 2).value == 3.14
        assert ws.cell(2, 3).value is True
        assert ws.cell(3, 2).value == 42
        assert ws.cell(3, 3).value is False

    def test_formatting_auto_filter(self, updater):
        result = updater.update_sheet(
            excel_binary=None,
            sheet_name="Filtered",
            headers=["A", "B"],
            data_rows=[["1", "2"]],
            formatting={"auto_filter": True},
        )

        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        ws = wb["Filtered"]
        assert ws.auto_filter.ref == "A1:B2"

    def test_formatting_freeze_header(self, updater):
        result = updater.update_sheet(
            excel_binary=None,
            sheet_name="Frozen",
            headers=["A"],
            data_rows=[["1"]],
            formatting={"freeze_header": True},
        )

        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        ws = wb["Frozen"]
        assert ws.freeze_panes == "A2"

    def test_empty_data_rows(self, updater):
        result = updater.update_sheet(
            excel_binary=None,
            sheet_name="Empty",
            headers=["H1", "H2"],
            data_rows=[],
        )

        assert result.rows_written == 0
        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        ws = wb["Empty"]
        assert ws.cell(1, 1).value == "H1"

    def test_large_dataset_performance(self, updater):
        """Verify 10,000 rows can be processed."""
        headers = [f"col_{i}" for i in range(10)]
        data_rows = [[f"val_{r}_{c}" for c in range(10)] for r in range(10_000)]

        result = updater.update_sheet(
            excel_binary=None,
            sheet_name="Large",
            headers=headers,
            data_rows=data_rows,
        )

        assert result.rows_written == 10_000
        assert len(result.xlsx_bytes) > 0
