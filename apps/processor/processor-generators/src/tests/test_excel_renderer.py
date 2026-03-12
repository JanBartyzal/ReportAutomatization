"""Tests for excel_renderer module."""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook, load_workbook

from src.generators.xls.service.excel_renderer import (
    ChartInput,
    ChartSeriesInput,
    ExcelRenderer,
    RenderResult,
    SheetConfig,
    TableInput,
)


def _make_template_with_text_placeholder(tag: str = "{{company_name}}") -> bytes:
    """Create a minimal XLSX template with a text placeholder."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = f"Company: {tag}"
    ws["B1"] = "Static content"
    ws["A2"] = "Date: {{report_date}}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _make_template_with_table_placeholder() -> bytes:
    """Create a XLSX template with a TABLE placeholder."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Report Title"
    ws["A3"] = "{{TABLE:data}}"
    ws["A10"] = "Footer content"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _make_template_with_chart_placeholder() -> bytes:
    """Create a XLSX template with a CHART placeholder."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Dashboard"
    ws["A3"] = "{{CHART:revenue}}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class TestPlaceholderSubstitution:
    """Test text placeholder substitution."""

    def test_single_placeholder(self):
        renderer = ExcelRenderer()
        template = _make_template_with_text_placeholder()
        result = renderer.render(
            template_bytes=template,
            placeholders={"company_name": "Acme Corp"},
            tables={},
            charts={},
        )

        assert isinstance(result, RenderResult)
        assert len(result.xlsx_bytes) > 0
        assert result.missing_placeholders == []

        # Verify the text was replaced
        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        ws = wb.active
        assert "Acme Corp" in str(ws["A1"].value)
        assert "{{company_name}}" not in str(ws["A1"].value)

    def test_multiple_placeholders(self):
        renderer = ExcelRenderer()
        template = _make_template_with_text_placeholder()
        result = renderer.render(
            template_bytes=template,
            placeholders={
                "company_name": "Acme Corp",
                "report_date": "2026-03-12",
            },
            tables={},
            charts={},
        )

        assert result.missing_placeholders == []

        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        ws = wb.active
        assert "Acme Corp" in str(ws["A1"].value)
        assert "2026-03-12" in str(ws["A2"].value)

    def test_missing_placeholder_reported(self):
        renderer = ExcelRenderer()
        template = _make_template_with_text_placeholder()
        result = renderer.render(
            template_bytes=template,
            placeholders={},  # No data provided
            tables={},
            charts={},
        )

        assert len(result.missing_placeholders) >= 1
        assert "{{company_name}}" in result.missing_placeholders


class TestTableFill:
    """Test table data population."""

    def test_table_fill(self):
        renderer = ExcelRenderer()
        template = _make_template_with_table_placeholder()
        result = renderer.render(
            template_bytes=template,
            placeholders={},
            tables={
                "data": TableInput(
                    headers=["Name", "Value", "Status"],
                    rows=[
                        ["Item A", "100", "Active"],
                        ["Item B", "200", "Inactive"],
                        ["Item C", "300", "Active"],
                    ],
                ),
            },
            charts={},
        )

        assert len(result.xlsx_bytes) > 0

        # Verify table data was written
        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        ws = wb.active
        # Headers should be at row 3 (where the placeholder was)
        assert ws.cell(row=3, column=1).value == "Name"
        assert ws.cell(row=3, column=2).value == "Value"
        assert ws.cell(row=3, column=3).value == "Status"
        # First data row at row 4
        assert ws.cell(row=4, column=1).value == "Item A"
        assert ws.cell(row=4, column=2).value == "100"

    def test_missing_table_reported(self):
        renderer = ExcelRenderer()
        template = _make_template_with_table_placeholder()
        result = renderer.render(
            template_bytes=template,
            placeholders={},
            tables={},  # No table data provided
            charts={},
        )

        assert "{{TABLE:data}}" in result.missing_placeholders


class TestMultiSheet:
    """Test multiple sheet creation."""

    def test_additional_data_sheets(self):
        renderer = ExcelRenderer()
        template = _make_template_with_text_placeholder()
        result = renderer.render(
            template_bytes=template,
            placeholders={"company_name": "Test Co"},
            tables={
                "summary": TableInput(
                    headers=["Metric", "Value"],
                    rows=[
                        ["Revenue", "1000"],
                        ["Costs", "800"],
                    ],
                ),
            },
            charts={},
            sheets=[
                SheetConfig(name="Summary Data", table_key="summary"),
            ],
        )

        assert len(result.xlsx_bytes) > 0

        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        assert "Summary Data" in wb.sheetnames
        ws = wb["Summary Data"]
        assert ws.cell(row=1, column=1).value == "Metric"
        assert ws.cell(row=1, column=2).value == "Value"
        assert ws.cell(row=2, column=1).value == "Revenue"
        assert ws.cell(row=3, column=1).value == "Costs"


class TestNoTemplate:
    """Test workbook creation from scratch (no template)."""

    def test_create_from_scratch(self):
        renderer = ExcelRenderer()
        result = renderer.render(
            template_bytes=None,
            placeholders={},
            tables={
                "inventory": TableInput(
                    headers=["SKU", "Quantity"],
                    rows=[
                        ["SKU-001", "50"],
                        ["SKU-002", "120"],
                    ],
                ),
            },
            charts={},
            sheets=[
                SheetConfig(name="Inventory", table_key="inventory"),
            ],
        )

        assert len(result.xlsx_bytes) > 0

        # Should be a valid XLSX
        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        assert "Inventory" in wb.sheetnames
        ws = wb["Inventory"]
        assert ws.cell(row=1, column=1).value == "SKU"
        assert ws.cell(row=2, column=1).value == "SKU-001"

    def test_output_is_valid_xlsx(self):
        renderer = ExcelRenderer()
        result = renderer.render(
            template_bytes=None,
            placeholders={},
            tables={},
            charts={},
        )

        # Should be parseable as a valid XLSX
        wb = load_workbook(io.BytesIO(result.xlsx_bytes))
        assert len(wb.sheetnames) >= 1
