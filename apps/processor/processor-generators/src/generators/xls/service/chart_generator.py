"""Generate charts in Excel worksheets using openpyxl.chart.

Handles ``{{CHART:metric_name}}`` placeholders by creating native Excel
charts at the placeholder cell position.
"""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING

from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from src.generators.xls.service.excel_renderer import ChartInput

logger = logging.getLogger(__name__)


class ChartGenerator:
    """Creates native Excel charts from chart input data."""

    def add_chart(
        self,
        ws: Worksheet,
        chart_data: ChartInput,
        anchor_row: int,
        anchor_col: int,
    ) -> None:
        """Create a chart and add it to the worksheet at the specified position.

        Args:
            ws: The target worksheet.
            chart_data: Chart specification including type, labels, and series.
            anchor_row: Row number where the chart should be anchored.
            anchor_col: Column number where the chart should be anchored.
        """
        chart_type = chart_data.chart_type.upper()

        if chart_type == "BAR":
            chart = self._create_bar_chart(ws, chart_data)
        elif chart_type == "LINE":
            chart = self._create_line_chart(ws, chart_data)
        elif chart_type == "PIE":
            chart = self._create_pie_chart(ws, chart_data)
        else:
            logger.warning("Unsupported chart type: %s", chart_type)
            return

        # Position the chart at the anchor cell
        anchor_cell = f"{get_column_letter(anchor_col)}{anchor_row}"
        ws.add_chart(chart, anchor_cell)

        logger.info(
            "Added %s chart at %s (%d series)",
            chart_type,
            anchor_cell,
            len(chart_data.series),
        )

    def _create_bar_chart(self, ws: Worksheet, chart_data: ChartInput) -> BarChart:
        """Create a bar chart with data series."""
        chart = BarChart()
        chart.type = "col"
        chart.title = chart_data.title or None
        chart.style = 10

        data_start_row, data_start_col = self._write_chart_data(ws, chart_data)

        num_series = len(chart_data.series)
        num_labels = len(chart_data.labels)

        # Categories (labels) reference
        categories = Reference(
            ws,
            min_col=data_start_col,
            min_row=data_start_row + 1,
            max_row=data_start_row + num_labels,
        )

        # Data series references
        for series_idx in range(num_series):
            data_ref = Reference(
                ws,
                min_col=data_start_col + 1 + series_idx,
                min_row=data_start_row,
                max_row=data_start_row + num_labels,
            )
            chart.add_data(data_ref, titles_from_data=True)

        chart.set_categories(categories)
        return chart

    def _create_line_chart(self, ws: Worksheet, chart_data: ChartInput) -> LineChart:
        """Create a line chart with data series."""
        chart = LineChart()
        chart.title = chart_data.title or None
        chart.style = 10

        data_start_row, data_start_col = self._write_chart_data(ws, chart_data)

        num_series = len(chart_data.series)
        num_labels = len(chart_data.labels)

        # Categories (labels) reference
        categories = Reference(
            ws,
            min_col=data_start_col,
            min_row=data_start_row + 1,
            max_row=data_start_row + num_labels,
        )

        # Data series references
        for series_idx in range(num_series):
            data_ref = Reference(
                ws,
                min_col=data_start_col + 1 + series_idx,
                min_row=data_start_row,
                max_row=data_start_row + num_labels,
            )
            chart.add_data(data_ref, titles_from_data=True)

        chart.set_categories(categories)
        return chart

    def _create_pie_chart(self, ws: Worksheet, chart_data: ChartInput) -> PieChart:
        """Create a pie chart using the first data series."""
        chart = PieChart()
        chart.title = chart_data.title or None
        chart.style = 10

        data_start_row, data_start_col = self._write_chart_data(ws, chart_data)

        num_labels = len(chart_data.labels)

        # Categories (labels) reference
        categories = Reference(
            ws,
            min_col=data_start_col,
            min_row=data_start_row + 1,
            max_row=data_start_row + num_labels,
        )

        # Data reference (first series only for pie charts)
        if chart_data.series:
            data_ref = Reference(
                ws,
                min_col=data_start_col + 1,
                min_row=data_start_row,
                max_row=data_start_row + num_labels,
            )
            chart.add_data(data_ref, titles_from_data=True)

        chart.set_categories(categories)
        return chart

    def _write_chart_data(
        self,
        ws: Worksheet,
        chart_data: ChartInput,
    ) -> tuple[int, int]:
        """Write chart data into a hidden area of the worksheet.

        Data is written starting at column 20 (T) to keep it away from
        visible content. Each chart call uses a progressively higher column
        to avoid collisions.

        Returns:
            Tuple of (start_row, start_col) where the data was written.
        """
        # Find a safe column offset to avoid collisions with existing data
        data_start_col = max(ws.max_column or 1, 1) + 2
        data_start_row = 1

        # Write header row: empty cell for labels column, then series names
        ws.cell(row=data_start_row, column=data_start_col).value = ""
        for series_idx, series in enumerate(chart_data.series):
            ws.cell(
                row=data_start_row,
                column=data_start_col + 1 + series_idx,
            ).value = series.name

        # Write labels and data rows
        for label_idx, label in enumerate(chart_data.labels):
            row = data_start_row + 1 + label_idx
            ws.cell(row=row, column=data_start_col).value = label
            for series_idx, series in enumerate(chart_data.series):
                value = series.values[label_idx] if label_idx < len(series.values) else 0
                ws.cell(row=row, column=data_start_col + 1 + series_idx).value = value

        return data_start_row, data_start_col
