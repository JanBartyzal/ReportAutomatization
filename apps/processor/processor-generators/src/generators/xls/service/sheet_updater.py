"""Partial sheet update engine for Excel workbooks.

Updates a single named sheet in an existing Excel workbook while preserving
all other sheets, their content, formatting, charts, and pivot tables.
"""

from __future__ import annotations

import io
import logging
import os
from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_excel_max_size_mb = int(os.environ.get("EXCEL_MAX_SIZE_MB", "50"))
MAX_INPUT_SIZE_BYTES = _excel_max_size_mb * 1024 * 1024
MAX_ROWS = 1_048_576
MAX_COLUMNS = 16_384


@dataclass
class SheetUpdateResult:
    """Result of a sheet update operation."""
    xlsx_bytes: bytes
    rows_written: int
    sheet_name: str


class SheetUpdater:
    """Updates a single sheet in an Excel workbook, preserving all other sheets."""

    def update_sheet(
        self,
        excel_binary: bytes | None,
        sheet_name: str,
        headers: list[str],
        data_rows: list[list],
        formatting: dict | None = None,
    ) -> SheetUpdateResult:
        """Update or create a sheet in an Excel workbook.

        Args:
            excel_binary: Existing Excel file bytes, or None/empty to create new.
            sheet_name: Target sheet name to overwrite or create.
            headers: Column header strings for row 1.
            data_rows: Data rows (list of lists of typed values).
            formatting: Optional dict with keys: auto_filter, freeze_header, auto_column_width.

        Returns:
            SheetUpdateResult with updated workbook bytes and metadata.
        """
        if formatting is None:
            formatting = {}

        if excel_binary and len(excel_binary) > MAX_INPUT_SIZE_BYTES:
            raise ValueError(f"Input Excel exceeds maximum size of {MAX_INPUT_SIZE_BYTES // (1024*1024)} MB")

        if len(data_rows) > MAX_ROWS:
            raise ValueError(f"Data exceeds Excel maximum of {MAX_ROWS} rows")

        if headers and len(headers) > MAX_COLUMNS:
            raise ValueError(f"Headers exceed Excel maximum of {MAX_COLUMNS} columns")

        # Load existing workbook or create new
        if excel_binary:
            wb = load_workbook(io.BytesIO(excel_binary))
        else:
            wb = Workbook()
            # Remove default sheet if we're creating a fresh workbook
            if wb.sheetnames == ["Sheet"] and sheet_name != "Sheet":
                del wb["Sheet"]

        # Get or create target sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._clear_sheet(ws)
            logger.info("Cleared existing sheet '%s'", sheet_name)
        else:
            ws = wb.create_sheet(title=sheet_name)
            logger.info("Created new sheet '%s'", sheet_name)

        # Write headers to row 1
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data rows starting from row 2
        rows_written = 0
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                self._set_cell_value(cell, value)
            rows_written += 1

        # Apply formatting
        if headers:
            self._apply_formatting(ws, headers, rows_written, formatting)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(
            "Sheet update complete: sheet='%s', headers=%d, rows=%d",
            sheet_name, len(headers), rows_written,
        )

        return SheetUpdateResult(
            xlsx_bytes=output.getvalue(),
            rows_written=rows_written,
            sheet_name=sheet_name,
        )

    def _clear_sheet(self, ws: Worksheet) -> None:
        """Clear all data from a worksheet without deleting it."""
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

        # Reset dimensions
        ws.delete_rows(1, ws.max_row)

    def _set_cell_value(self, cell, value) -> None:
        """Set cell value with appropriate type handling."""
        if value is None:
            cell.value = None
        elif isinstance(value, bool):
            cell.value = value
        elif isinstance(value, (int, float)):
            cell.value = value
        elif isinstance(value, str):
            # Try to detect date strings in ISO 8601 format
            cell.value = value
        elif isinstance(value, datetime):
            cell.value = value
            cell.number_format = "YYYY-MM-DD"
        else:
            cell.value = str(value)

    def _apply_formatting(
        self, ws: Worksheet, headers: list[str],
        rows_written: int, formatting: dict
    ) -> None:
        """Apply formatting options to the sheet."""
        num_cols = len(headers)
        last_row = rows_written + 1  # +1 for header row

        # Auto-filter on headers
        if formatting.get("auto_filter", False) and num_cols > 0:
            last_col_letter = get_column_letter(num_cols)
            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

        # Freeze header row
        if formatting.get("freeze_header", False):
            ws.freeze_panes = "A2"

        # Auto column width
        if formatting.get("auto_column_width", False):
            for col_idx in range(1, num_cols + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                        min_row=1, max_row=last_row):
                    for cell in row:
                        if cell.value is not None:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length:
                                max_length = cell_len
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
