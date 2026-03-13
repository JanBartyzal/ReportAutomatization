"""Core Excel rendering engine.

Orchestrates placeholder substitution, table population, and chart insertion
into an Excel template (or a new workbook) while preserving original formatting.

Supports three placeholder types:
- ``{{variable_name}}`` -- text replacement
- ``{{TABLE:table_name}}`` -- table data population
- ``{{CHART:chart_name}}`` -- chart insertion
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_PLACEHOLDER_RE = re.compile(r"\{\{(TABLE:|CHART:)?([^}]+)\}\}")
_TABLE_PLACEHOLDER_RE = re.compile(r"\{\{TABLE:([^}]+)\}\}")
_CHART_PLACEHOLDER_RE = re.compile(r"\{\{CHART:([^}]+)\}\}")


@dataclass
class TableInput:
    """Table data for a single placeholder."""
    headers: list[str]
    rows: list[list[str]]


@dataclass
class ChartInput:
    """Chart data for a single placeholder."""
    chart_type: str  # BAR, LINE, PIE
    title: str = ""
    labels: list[str] = field(default_factory=list)
    series: list[ChartSeriesInput] = field(default_factory=list)


@dataclass
class ChartSeriesInput:
    """A single data series for a chart."""
    name: str
    values: list[float]


@dataclass
class SheetConfig:
    """Configuration for an additional data sheet."""
    name: str
    table_key: str


@dataclass
class RenderResult:
    """Result of rendering an Excel template."""
    xlsx_bytes: bytes
    missing_placeholders: list[str]


class ExcelRenderer:
    """Renders Excel reports from templates with placeholder substitution, tables, and charts."""

    def render(
        self,
        template_bytes: bytes | None,
        placeholders: dict[str, str],
        tables: dict[str, TableInput],
        charts: dict[str, ChartInput],
        sheets: list[SheetConfig] | None = None,
    ) -> RenderResult:
        """Render an Excel report.

        Args:
            template_bytes: Raw XLSX template bytes, or None to create a new workbook.
            placeholders: Map of placeholder key to replacement text value.
            tables: Map of table placeholder key to ``TableInput``.
            charts: Map of chart placeholder key to ``ChartInput``.
            sheets: Optional list of additional data sheets to create.

        Returns:
            ``RenderResult`` with generated XLSX bytes and list of missing placeholders.
        """
        missing: list[str] = []

        if template_bytes is not None:
            wb = load_workbook(io.BytesIO(template_bytes))
        else:
            wb = Workbook()

        # Process each worksheet in the workbook
        for ws in wb.worksheets:
            self._process_worksheet(ws, placeholders, tables, charts, missing)

        # Add additional data sheets if requested
        if sheets:
            for sheet_config in sheets:
                self._add_data_sheet(wb, sheet_config, tables)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(
            "Render complete: %d placeholders, %d tables, %d charts, %d missing",
            len(placeholders),
            len(tables),
            len(charts),
            len(missing),
        )

        return RenderResult(xlsx_bytes=output.getvalue(), missing_placeholders=missing)

    def _process_worksheet(
        self,
        ws: Worksheet,
        placeholders: dict[str, str],
        tables: dict[str, TableInput],
        charts: dict[str, ChartInput],
        missing: list[str],
    ) -> None:
        """Process a single worksheet: substitute placeholders, fill tables, add charts."""
        # First pass: substitute text placeholders
        self._substitute_placeholders(ws, placeholders, missing)

        # Second pass: fill tables (iterate in reverse to avoid row-shift issues)
        table_cells = self._find_table_placeholders(ws)
        for row_idx, col_idx, key in reversed(table_cells):
            if key in tables:
                self._fill_table(ws, row_idx, col_idx, tables[key])
            else:
                ws.cell(row=row_idx, column=col_idx).value = f"[MISSING TABLE: {key}]"
                missing.append(f"{{{{TABLE:{key}}}}}")

        # Third pass: insert charts
        chart_cells = self._find_chart_placeholders(ws)
        for row_idx, col_idx, key in chart_cells:
            if key in charts:
                from src.generators.xls.service.chart_generator import ChartGenerator

                chart_gen = ChartGenerator()
                chart_gen.add_chart(ws, charts[key], row_idx, col_idx)
                # Clear the placeholder cell
                ws.cell(row=row_idx, column=col_idx).value = None
            else:
                ws.cell(row=row_idx, column=col_idx).value = f"[MISSING CHART: {key}]"
                missing.append(f"{{{{CHART:{key}}}}}")

    def _substitute_placeholders(
        self,
        ws: Worksheet,
        placeholders: dict[str, str],
        missing: list[str],
    ) -> None:
        """Iterate cells and replace ``{{var_name}}`` placeholders with values."""
        # Determine if placeholders dict was explicitly provided (non-empty)
        # If empty, we report ALL placeholders as missing
        # If non-empty, we only report missing for keys that were explicitly provided but have no value
        placeholders_provided = bool(placeholders)
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue

                original = cell.value
                # Skip TABLE: and CHART: placeholders (handled separately)
                if _TABLE_PLACEHOLDER_RE.search(original) or _CHART_PLACEHOLDER_RE.search(original):
                    continue

                # Find all text placeholders in this cell
                new_value = original
                for match in _PLACEHOLDER_RE.finditer(original):
                    prefix = match.group(1) or ""
                    if prefix:
                        continue  # Skip TABLE:/CHART: prefixed placeholders
                    key = match.group(2).strip()
                    tag = match.group(0)
                    
                    if key in placeholders:
                        # Key explicitly provided - substitute if value is not None/empty
                        if placeholders[key] is not None and placeholders[key] != "":
                            new_value = new_value.replace(tag, placeholders[key])
                        else:
                            # Key provided but value is empty - report as missing
                            missing.append(tag)
                    elif not placeholders_provided:
                        # No placeholders provided at all - report all as missing
                        missing.append(tag)
                    # If key not in placeholders and placeholders dict is non-empty, 
                    # don't report as missing (user may not know about this placeholder)

                if new_value != original:
                    cell.value = new_value

    def _find_table_placeholders(self, ws: Worksheet) -> list[tuple[int, int, str]]:
        """Find all ``{{TABLE:name}}`` cells and return (row, col, key) tuples."""
        results: list[tuple[int, int, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                match = _TABLE_PLACEHOLDER_RE.search(cell.value)
                if match:
                    results.append((cell.row, cell.column, match.group(1).strip()))
        return results

    def _find_chart_placeholders(self, ws: Worksheet) -> list[tuple[int, int, str]]:
        """Find all ``{{CHART:name}}`` cells and return (row, col, key) tuples."""
        results: list[tuple[int, int, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                match = _CHART_PLACEHOLDER_RE.search(cell.value)
                if match:
                    results.append((cell.row, cell.column, match.group(1).strip()))
        return results

    def _fill_table(
        self,
        ws: Worksheet,
        start_row: int,
        start_col: int,
        table_data: TableInput,
    ) -> None:
        """Fill table data starting at the placeholder cell.

        Inserts headers at the placeholder row, then data rows below it.
        Rows are inserted to avoid overwriting existing content.
        """
        # Insert blank rows below the placeholder to make room (excluding the placeholder row itself)
        if len(table_data.rows) > 0:
            ws.insert_rows(start_row + 1, amount=len(table_data.rows))

        # Write headers at the placeholder row
        for col_offset, header in enumerate(table_data.headers):
            ws.cell(row=start_row, column=start_col + col_offset).value = header

        # Write data rows
        for row_offset, row_data in enumerate(table_data.rows):
            for col_offset, value in enumerate(row_data):
                ws.cell(
                    row=start_row + 1 + row_offset,
                    column=start_col + col_offset,
                ).value = value

        logger.info(
            "Filled table at row=%d col=%d: %d headers, %d data rows",
            start_row,
            start_col,
            len(table_data.headers),
            len(table_data.rows),
        )

    def _add_data_sheet(
        self,
        wb: Workbook,
        sheet_config: SheetConfig,
        tables: dict[str, TableInput],
    ) -> None:
        """Create a new worksheet with table data."""
        if sheet_config.table_key not in tables:
            logger.warning(
                "Table key '%s' not found for sheet '%s'",
                sheet_config.table_key,
                sheet_config.name,
            )
            return

        table_data = tables[sheet_config.table_key]
        ws = wb.create_sheet(title=sheet_config.name)

        # Write headers in row 1
        for col_idx, header in enumerate(table_data.headers, start=1):
            ws.cell(row=1, column=col_idx).value = header

        # Write data rows starting from row 2
        for row_idx, row_data in enumerate(table_data.rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx).value = value

        logger.info(
            "Created data sheet '%s' with %d headers and %d rows",
            sheet_config.name,
            len(table_data.headers),
            len(table_data.rows),
        )
