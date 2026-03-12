"""Integration tests for ExcelParser + DataTypeDetector pipeline."""

from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest

from src.atomizers.xls.service.data_type_detector import DataTypeDetector
from src.atomizers.xls.service.excel_parser import ExcelParser


def _create_test_workbook(
    sheets: dict[str, tuple[list[str], list[list[str | int | float | None]]]] | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    default_sheet = wb.active

    if sheets is None:
        sheets = {"Sheet1": (["Name", "Value"], [["Alpha", 100], ["Beta", 200]])}

    if sheets:
        wb.remove(default_sheet)
        for sheet_name, (headers, rows) in sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            for row_idx, row_data in enumerate(rows, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class TestExcelParserIntegration:
    def test_full_extraction_pipeline(self) -> None:
        excel_bytes = _create_test_workbook({
            "Data": (
                ["Name", "Amount", "Date", "Rate"],
                [
                    ["Widget", 100, "2024-01-15", "10%"],
                    ["Gadget", 200, "2024-02-28", "20%"],
                    ["Doohickey", 300, "2024-03-31", "30%"],
                ],
            ),
        })

        parser = ExcelParser()
        detector = DataTypeDetector()

        wb = parser.open(excel_bytes)
        structure = parser.extract_structure(wb)

        assert len(structure.sheets) == 1
        assert structure.sheets[0].name == "Data"

        content = parser.extract_sheet_content(wb, 0)
        assert content.headers == ["Name", "Amount", "Date", "Rate"]
        assert len(content.rows) == 3

        for col_idx, header in enumerate(content.headers):
            col_values = [r.cells[col_idx] for r in content.rows]
            detected = detector.detect_column_type(col_values)

            if header == "Name":
                assert detected == "STRING"
            elif header == "Amount":
                assert detected == "NUMBER"
            elif header == "Date":
                assert detected == "DATE"
            elif header == "Rate":
                assert detected == "PERCENTAGE"

        wb.close()

    def test_multi_sheet_extraction(self) -> None:
        excel_bytes = _create_test_workbook({
            "Sales": (["Product", "Revenue"], [["A", 1000], ["B", 2000]]),
            "Costs": (["Item", "Amount"], [["Rent", 500]]),
        })

        parser = ExcelParser()
        wb = parser.open(excel_bytes)
        structure = parser.extract_structure(wb)

        assert len(structure.sheets) == 2

        for sheet_meta in structure.sheets:
            content = parser.extract_sheet_content(wb, sheet_meta.sheet_index)
            assert content.sheet_name == sheet_meta.name
            assert len(content.headers) > 0

        wb.close()

    def test_extraction_with_errors_per_sheet(self) -> None:
        excel_bytes = _create_test_workbook({"Good": (["A"], [["1"]])})

        parser = ExcelParser()
        wb = parser.open(excel_bytes)

        content = parser.extract_sheet_content(wb, 0)
        assert content.sheet_name == "Good"

        with pytest.raises(IndexError):
            parser.extract_sheet_content(wb, 99)

        wb.close()

    def test_empty_sheet_extraction(self) -> None:
        wb_obj = openpyxl.Workbook()
        ws = wb_obj.active
        ws.title = "Empty"
        buf = BytesIO()
        wb_obj.save(buf)
        buf.seek(0)
        excel_bytes = buf.read()

        parser = ExcelParser()
        wb = parser.open(excel_bytes)
        content = parser.extract_sheet_content(wb, 0)
        wb.close()

        assert content.headers == []
        assert content.rows == []
