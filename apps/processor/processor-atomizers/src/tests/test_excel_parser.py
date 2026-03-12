"""Unit tests for ExcelParser using programmatically created Excel files."""

from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest

from src.atomizers.xls.service.excel_parser import ExcelParser


# ---------------------------------------------------------------------------
# Fixtures -- create Excel workbooks in memory
# ---------------------------------------------------------------------------


def _create_simple_workbook(
    headers: list[str],
    rows: list[list[str | int | float | None]],
    sheet_name: str = "Sheet1",
) -> bytes:
    """Create a simple single-sheet workbook and return as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _create_multi_sheet_workbook(
    sheets: dict[str, tuple[list[str], list[list[str | int | float | None]]]],
) -> bytes:
    """Create a multi-sheet workbook and return as bytes."""
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _create_merged_cells_workbook() -> bytes:
    """Create a workbook with merged cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"

    ws.cell(row=1, column=1, value="Category")
    ws.cell(row=1, column=2, value="Value")

    ws.cell(row=2, column=1, value="Group A")
    ws.cell(row=2, column=2, value=100)
    ws.cell(row=3, column=1, value="")
    ws.cell(row=3, column=2, value=200)

    ws.merge_cells("A2:A3")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _create_hidden_sheet_workbook() -> bytes:
    """Create a workbook with a hidden sheet."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Visible"
    ws1.cell(row=1, column=1, value="Header")
    ws1.cell(row=2, column=1, value="Data")

    ws2 = wb.create_sheet(title="Hidden")
    ws2.sheet_state = "hidden"
    ws2.cell(row=1, column=1, value="Secret")
    ws2.cell(row=2, column=1, value="Value")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _create_empty_rows_workbook() -> bytes:
    """Create a workbook with some empty rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sparse"

    ws.cell(row=1, column=1, value="Name")
    ws.cell(row=1, column=2, value="Age")

    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=2, column=2, value=30)

    ws.cell(row=4, column=1, value="Bob")
    ws.cell(row=4, column=2, value=25)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Tests - Structure Extraction
# ---------------------------------------------------------------------------


class TestExtractStructure:
    """Tests for ExcelParser.extract_structure."""

    def test_single_sheet_structure(self) -> None:
        data = _create_simple_workbook(
            headers=["A", "B", "C"],
            rows=[["1", "2", "3"], ["4", "5", "6"]],
        )
        parser = ExcelParser()
        wb = parser.open(data)
        structure = parser.extract_structure(wb)
        wb.close()

        assert len(structure.sheets) == 1
        sheet = structure.sheets[0]
        assert sheet.sheet_index == 0
        assert sheet.name == "Sheet1"
        assert sheet.row_count == 3
        assert sheet.col_count == 3
        assert sheet.has_merged_cells is False
        assert sheet.is_hidden is False

    def test_multi_sheet_structure(self) -> None:
        data = _create_multi_sheet_workbook({
            "Sales": (["Product", "Amount"], [["Widget", 100]]),
            "Costs": (["Item", "Cost"], [["Rent", 500], ["Salary", 3000]]),
        })
        parser = ExcelParser()
        wb = parser.open(data)
        structure = parser.extract_structure(wb)
        wb.close()

        assert len(structure.sheets) == 2
        assert structure.sheets[0].name == "Sales"
        assert structure.sheets[1].name == "Costs"

    def test_merged_cells_detected(self) -> None:
        data = _create_merged_cells_workbook()
        parser = ExcelParser()
        wb = parser.open(data)
        structure = parser.extract_structure(wb)
        wb.close()

        assert structure.sheets[0].has_merged_cells is True

    def test_hidden_sheet_detected(self) -> None:
        data = _create_hidden_sheet_workbook()
        parser = ExcelParser()
        wb = parser.open(data)
        structure = parser.extract_structure(wb)
        wb.close()

        assert len(structure.sheets) == 2
        visible = [s for s in structure.sheets if not s.is_hidden]
        hidden = [s for s in structure.sheets if s.is_hidden]
        assert len(visible) == 1
        assert len(hidden) == 1
        assert hidden[0].name == "Hidden"

    def test_empty_workbook(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = buf.read()

        parser = ExcelParser()
        wb2 = parser.open(data)
        structure = parser.extract_structure(wb2)
        wb2.close()

        assert len(structure.sheets) == 1
        assert structure.sheets[0].name == "Empty"


# ---------------------------------------------------------------------------
# Tests - Sheet Content Extraction
# ---------------------------------------------------------------------------


class TestExtractSheetContent:
    """Tests for ExcelParser.extract_sheet_content."""

    def test_basic_content_extraction(self) -> None:
        data = _create_simple_workbook(
            headers=["Name", "Value"],
            rows=[["Alpha", 100], ["Beta", 200]],
        )
        parser = ExcelParser()
        wb = parser.open(data)
        content = parser.extract_sheet_content(wb, 0)
        wb.close()

        assert content.sheet_index == 0
        assert content.sheet_name == "Sheet1"
        assert content.headers == ["Name", "Value"]
        assert len(content.rows) == 2
        assert content.rows[0].cells == ["Alpha", "100"]
        assert content.rows[1].cells == ["Beta", "200"]

    def test_merged_cells_are_filled(self) -> None:
        data = _create_merged_cells_workbook()
        parser = ExcelParser()
        wb = parser.open(data)
        content = parser.extract_sheet_content(wb, 0)
        wb.close()

        assert content.rows[0].cells[0] == "Group A"
        assert content.rows[1].cells[0] == "Group A"

    def test_empty_rows_skipped(self) -> None:
        data = _create_empty_rows_workbook()
        parser = ExcelParser(empty_row_threshold=0)
        wb = parser.open(data)
        content = parser.extract_sheet_content(wb, 0)
        wb.close()

        non_empty_rows = [r for r in content.rows if any(c.strip() for c in r.cells)]
        assert len(non_empty_rows) == 2

    def test_out_of_range_raises(self) -> None:
        data = _create_simple_workbook(headers=["A"], rows=[["1"]])
        parser = ExcelParser()
        wb = parser.open(data)

        with pytest.raises(IndexError):
            parser.extract_sheet_content(wb, 5)

        wb.close()

    def test_negative_index_raises(self) -> None:
        data = _create_simple_workbook(headers=["A"], rows=[["1"]])
        parser = ExcelParser()
        wb = parser.open(data)

        with pytest.raises(IndexError):
            parser.extract_sheet_content(wb, -1)

        wb.close()

    def test_none_values_become_empty_string(self) -> None:
        data = _create_simple_workbook(
            headers=["A", "B"],
            rows=[[None, "val"], ["val", None]],
        )
        parser = ExcelParser()
        wb = parser.open(data)
        content = parser.extract_sheet_content(wb, 0)
        wb.close()

        assert content.rows[0].cells[0] == ""
        assert content.rows[1].cells[1] == ""

    def test_boolean_values(self) -> None:
        data = _create_simple_workbook(
            headers=["Flag"],
            rows=[[True], [False]],
        )
        parser = ExcelParser()
        wb = parser.open(data)
        content = parser.extract_sheet_content(wb, 0)
        wb.close()

        assert content.rows[0].cells[0] == "TRUE"
        assert content.rows[1].cells[0] == "FALSE"

    def test_float_whole_number_no_decimal(self) -> None:
        data = _create_simple_workbook(
            headers=["Amount"],
            rows=[[100.0], [200.5]],
        )
        parser = ExcelParser()
        wb = parser.open(data)
        content = parser.extract_sheet_content(wb, 0)
        wb.close()

        assert content.rows[0].cells[0] == "100"
        assert content.rows[1].cells[0] == "200.5"

    def test_multi_sheet_content_extraction(self) -> None:
        data = _create_multi_sheet_workbook({
            "First": (["X"], [["a"]]),
            "Second": (["Y"], [["b"]]),
        })
        parser = ExcelParser()
        wb = parser.open(data)

        content0 = parser.extract_sheet_content(wb, 0)
        content1 = parser.extract_sheet_content(wb, 1)
        wb.close()

        assert content0.sheet_name == "First"
        assert content0.headers == ["X"]
        assert content1.sheet_name == "Second"
        assert content1.headers == ["Y"]
