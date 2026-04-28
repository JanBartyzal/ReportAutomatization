"""Parser for ServiceNow export files (CSV, JSON, Excel).

Supports three export formats produced by ServiceNow:
- CSV  – Table API export (``/api/now/table/<name>.csv``)
- JSON – Table API export (``Accept: application/json``)
- EXCEL – ServiceNow report export (.xlsx)

All three formats are normalised into the same
:class:`ServiceNowParsingResult` that mirrors the Excel atomizer output,
so the existing Sink pipeline requires no changes.
"""

from __future__ import annotations

import csv
import io
import json
import logging
from typing import Any

from src.atomizers.servicenow.models.context import (
    ServiceNowExportFormat,
    ServiceNowParsingResult,
    ServiceNowStructureData,
    ServiceNowTableMetadata,
    ServiceNowTableRow,
)

logger = logging.getLogger(__name__)

# ServiceNow system columns that carry no analytical value
_SNOW_SYSTEM_COLUMNS: frozenset[str] = frozenset(
    {
        "sys_id",
        "sys_created_by",
        "sys_created_on",
        "sys_mod_count",
        "sys_updated_by",
        "sys_updated_on",
        "sys_tags",
    }
)


class ServiceNowParser:
    """Parses ServiceNow exports into a normalised tabular representation."""

    def __init__(
        self,
        strip_system_columns: bool = True,
        max_rows: int = 100_000,
    ) -> None:
        self._strip_system = strip_system_columns
        self._max_rows = max_rows

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def detect_format(self, raw: bytes, filename: str = "") -> ServiceNowExportFormat:
        """Detect the export format from content and/or filename."""
        name = filename.lower()
        if name.endswith(".xlsx") or name.endswith(".xls"):
            return ServiceNowExportFormat.EXCEL

        # Try JSON
        try:
            obj = json.loads(raw[:4096].decode("utf-8", errors="ignore"))
            if isinstance(obj, dict) and "result" in obj:
                return ServiceNowExportFormat.JSON
        except (json.JSONDecodeError, ValueError):
            pass

        # Default to CSV
        return ServiceNowExportFormat.CSV

    def extract_structure(
        self,
        raw: bytes,
        fmt: ServiceNowExportFormat,
        snow_table_name: str = "",
    ) -> ServiceNowStructureData:
        """Return high-level structure without loading all rows."""
        if fmt == ServiceNowExportFormat.JSON:
            tables = self._json_structure(raw, snow_table_name)
        elif fmt == ServiceNowExportFormat.EXCEL:
            tables = self._excel_structure(raw, snow_table_name)
        else:
            tables = self._csv_structure(raw, snow_table_name)

        return ServiceNowStructureData(detected_format=fmt, tables=tables)

    def extract_table(
        self,
        raw: bytes,
        fmt: ServiceNowExportFormat,
        table_index: int = 0,
        snow_table_name: str = "",
    ) -> ServiceNowParsingResult:
        """Extract a single table (full rows + headers)."""
        if fmt == ServiceNowExportFormat.JSON:
            return self._parse_json(raw, table_index, snow_table_name)
        if fmt == ServiceNowExportFormat.EXCEL:
            return self._parse_excel(raw, table_index, snow_table_name)
        return self._parse_csv(raw, table_index, snow_table_name)

    # ------------------------------------------------------------------
    # CSV parsing
    # ------------------------------------------------------------------

    def _csv_structure(self, raw: bytes, snow_table_name: str) -> list[ServiceNowTableMetadata]:
        reader, headers = self._csv_reader(raw)
        row_count = sum(1 for _ in reader)
        return [ServiceNowTableMetadata(
            table_index=0,
            name=snow_table_name or "data",
            row_count=row_count,
            col_count=len(headers),
            snow_table_name=snow_table_name,
        )]

    def _parse_csv(
        self, raw: bytes, table_index: int, snow_table_name: str
    ) -> ServiceNowParsingResult:
        if table_index != 0:
            raise IndexError(f"CSV exports contain only one table (index 0); got {table_index}")

        reader, headers = self._csv_reader(raw)
        rows: list[ServiceNowTableRow] = []
        for i, row_dict in enumerate(reader):
            if i >= self._max_rows:
                logger.warning("CSV row limit (%d) reached for fileId; truncating", self._max_rows)
                break
            cells = [str(row_dict.get(h, "") or "") for h in headers]
            rows.append(ServiceNowTableRow(row_index=i, cells=cells))

        return ServiceNowParsingResult(
            table_index=0,
            table_name=snow_table_name or "data",
            headers=headers,
            rows=rows,
            snow_table_name=snow_table_name,
            detected_format=ServiceNowExportFormat.CSV,
        )

    def _csv_reader(self, raw: bytes) -> tuple[csv.DictReader, list[str]]:
        """Decode bytes, detect delimiter, return (DictReader, headers)."""
        text = raw.decode(self._detect_encoding(raw), errors="replace")
        # Detect delimiter by sniffing
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel  # type: ignore[assignment]

        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        headers = self._filter_headers(list(reader.fieldnames or []))
        return reader, headers

    # ------------------------------------------------------------------
    # JSON parsing
    # ------------------------------------------------------------------

    def _json_structure(self, raw: bytes, snow_table_name: str) -> list[ServiceNowTableMetadata]:
        payload = self._load_json(raw)
        records = payload.get("result", [])
        headers = list(records[0].keys()) if records else []
        headers = self._filter_headers(headers)
        return [ServiceNowTableMetadata(
            table_index=0,
            name=snow_table_name or "result",
            row_count=len(records),
            col_count=len(headers),
            snow_table_name=snow_table_name,
        )]

    def _parse_json(
        self, raw: bytes, table_index: int, snow_table_name: str
    ) -> ServiceNowParsingResult:
        if table_index != 0:
            raise IndexError(f"JSON exports contain only one table (index 0); got {table_index}")

        payload = self._load_json(raw)
        records: list[dict[str, Any]] = payload.get("result", [])
        if not records:
            return ServiceNowParsingResult(
                table_index=0,
                table_name=snow_table_name or "result",
                headers=[],
                rows=[],
                snow_table_name=snow_table_name,
                detected_format=ServiceNowExportFormat.JSON,
            )

        headers = self._filter_headers(list(records[0].keys()))
        rows: list[ServiceNowTableRow] = []
        for i, rec in enumerate(records[: self._max_rows]):
            cells = [str(rec.get(h, "") or "") for h in headers]
            rows.append(ServiceNowTableRow(row_index=i, cells=cells))

        return ServiceNowParsingResult(
            table_index=0,
            table_name=snow_table_name or "result",
            headers=headers,
            rows=rows,
            snow_table_name=snow_table_name,
            detected_format=ServiceNowExportFormat.JSON,
        )

    def _load_json(self, raw: bytes) -> dict[str, Any]:
        text = raw.decode(self._detect_encoding(raw), errors="replace")
        return json.loads(text)

    # ------------------------------------------------------------------
    # Excel parsing (delegates to openpyxl, same as Excel atomizer)
    # ------------------------------------------------------------------

    def _excel_structure(self, raw: bytes, snow_table_name: str) -> list[ServiceNowTableMetadata]:
        import openpyxl  # type: ignore[import-untyped]

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        tables: list[ServiceNowTableMetadata] = []
        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            tables.append(ServiceNowTableMetadata(
                table_index=idx,
                name=sheet_name,
                row_count=ws.max_row or 0,
                col_count=ws.max_column or 0,
                snow_table_name=snow_table_name,
            ))
        wb.close()
        return tables

    def _parse_excel(
        self, raw: bytes, table_index: int, snow_table_name: str
    ) -> ServiceNowParsingResult:
        import openpyxl  # type: ignore[import-untyped]

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            if table_index >= len(wb.sheetnames):
                raise IndexError(
                    f"Sheet index {table_index} out of range ({len(wb.sheetnames)} sheets)"
                )
            sheet_name = wb.sheetnames[table_index]
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if not all_rows:
            return ServiceNowParsingResult(
                table_index=table_index,
                table_name=sheet_name,
                headers=[],
                rows=[],
                snow_table_name=snow_table_name,
                detected_format=ServiceNowExportFormat.EXCEL,
            )

        headers = self._filter_headers([str(c) if c is not None else "" for c in all_rows[0]])
        rows: list[ServiceNowTableRow] = []
        for i, raw_row in enumerate(all_rows[1: self._max_rows + 1]):
            cells = [str(c) if c is not None else "" for c in raw_row[: len(headers)]]
            rows.append(ServiceNowTableRow(row_index=i, cells=cells))

        return ServiceNowParsingResult(
            table_index=table_index,
            table_name=sheet_name,
            headers=headers,
            rows=rows,
            snow_table_name=snow_table_name,
            detected_format=ServiceNowExportFormat.EXCEL,
        )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _filter_headers(self, headers: list[str]) -> list[str]:
        if not self._strip_system:
            return headers
        return [h for h in headers if h.lower() not in _SNOW_SYSTEM_COLUMNS]

    @staticmethod
    def _detect_encoding(raw: bytes) -> str:
        try:
            import chardet  # type: ignore[import-untyped]
            result = chardet.detect(raw[:8192])
            return result.get("encoding") or "utf-8"
        except ImportError:
            return "utf-8"
