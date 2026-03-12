"""Core Excel parsing logic using openpyxl.

Extracts workbook structure, sheet content (headers + rows), and handles
merged cells, hidden sheets, formula cells (via data_only), and empty row
filtering.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data classes for parsed output
# ---------------------------------------------------------------------------


@dataclass(slots=True)
class SheetMetadataData:
    """Metadata about a single worksheet."""

    sheet_index: int
    name: str
    row_count: int
    col_count: int
    has_merged_cells: bool
    is_hidden: bool = False


@dataclass(slots=True)
class SheetRowData:
    """A single row of cell values."""

    row_index: int
    cells: list[str]


@dataclass(slots=True)
class SheetContentData:
    """All content extracted from a single sheet."""

    sheet_index: int
    sheet_name: str
    headers: list[str]
    rows: list[SheetRowData]


@dataclass(slots=True)
class WorkbookStructure:
    """Complete workbook structure."""

    sheets: list[SheetMetadataData]


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


class ExcelParser:
    """Stateless parser that wraps openpyxl operations.

    Usage::

        parser = ExcelParser()
        wb = parser.open(raw_bytes)
        structure = parser.extract_structure(wb)
        content = parser.extract_sheet_content(wb, sheet_index=0)
    """

    def __init__(self, empty_row_threshold: int = 0) -> None:
        self._empty_row_threshold = empty_row_threshold

    @staticmethod
    def open(source: bytes | BytesIO) -> openpyxl.Workbook:
        """Open an Excel file and return a ``Workbook`` object."""
        if isinstance(source, bytes):
            source = BytesIO(source)
        return openpyxl.load_workbook(source, data_only=True, read_only=False)

    def extract_structure(self, wb: openpyxl.Workbook) -> WorkbookStructure:
        """Extract the high-level structure of a workbook."""
        sheets_info: list[SheetMetadataData] = []

        for idx, sheet_name in enumerate(wb.sheetnames):
            ws: Worksheet = wb[sheet_name]

            row_count = ws.max_row or 0
            col_count = ws.max_column or 0
            has_merged = len(ws.merged_cells.ranges) > 0
            is_hidden = ws.sheet_state != "visible"

            sheets_info.append(
                SheetMetadataData(
                    sheet_index=idx,
                    name=sheet_name,
                    row_count=row_count,
                    col_count=col_count,
                    has_merged_cells=has_merged,
                    is_hidden=is_hidden,
                )
            )

        return WorkbookStructure(sheets=sheets_info)

    def extract_sheet_content(
        self,
        wb: openpyxl.Workbook,
        sheet_index: int,
    ) -> SheetContentData:
        """Extract all content from a specific sheet."""
        if sheet_index < 0 or sheet_index >= len(wb.sheetnames):
            raise IndexError(
                f"Sheet index {sheet_index} out of range (0..{len(wb.sheetnames) - 1})"
            )

        sheet_name = wb.sheetnames[sheet_index]
        ws: Worksheet = wb[sheet_name]

        self._unmerge_and_fill(ws)

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0 or max_col == 0:
            return SheetContentData(
                sheet_index=sheet_index,
                sheet_name=sheet_name,
                headers=[],
                rows=[],
            )

        headers: list[str] = []
        for col_idx in range(1, max_col + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            headers.append(self._cell_to_str(cell_value))

        rows: list[SheetRowData] = []
        for row_idx in range(2, max_row + 1):
            cells: list[str] = []
            for col_idx in range(1, max_col + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                cells.append(self._cell_to_str(cell_value))

            non_empty_count = sum(1 for c in cells if c.strip())
            if non_empty_count <= self._empty_row_threshold:
                continue

            rows.append(
                SheetRowData(
                    row_index=row_idx - 1,
                    cells=cells,
                )
            )

        return SheetContentData(
            sheet_index=sheet_index,
            sheet_name=sheet_name,
            headers=headers,
            rows=rows,
        )

    @staticmethod
    def _unmerge_and_fill(ws: Worksheet) -> None:
        """Unmerge all merged cell ranges and fill each cell with the top-left value."""
        merged_ranges = list(ws.merged_cells.ranges)

        for merged_range in merged_ranges:
            top_left_cell = ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            )
            fill_value = top_left_cell.value

            ws.unmerge_cells(str(merged_range))

            for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).value = fill_value

    @staticmethod
    def _cell_to_str(value: Any) -> str:
        """Convert a cell value to a string representation."""
        if value is None:
            return ""
        if isinstance(value, bool):
            return str(value).upper()
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            return str(value)
        return str(value)
